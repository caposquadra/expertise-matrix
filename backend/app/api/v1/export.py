import csv
import io

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.api.deps import get_manager_user
from app.core.database import get_db
from app.models import Assessment, Employee

router = APIRouter(prefix="/export", tags=["export"])

LEVEL_COLORS = {
    1: "FF6B6B",
    2: "FFA94D",
    3: "FFD43B",
    4: "69DB7C",
    5: "2F9E44",
}


@router.get("/csv")
async def export_csv(
    db: AsyncSession = Depends(get_db),
    _: Employee = Depends(get_manager_user),
):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(
        [
            "Сотрудник",
            "Email",
            "Навык",
            "Категория",
            "Самооценка",
            "Оценка руководителя",
            "Цель",
        ]
    )

    rows = await _fetch_matrix_data(db)
    for r in rows:
        writer.writerow(
            [
                r["employee_name"],
                r["email"],
                r["skill_name"],
                r["category"],
                r["self_level"],
                r["manager_level"],
                r["target_level"],
            ]
        )

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=expertise_matrix.csv"},
    )


@router.get("/excel")
async def export_excel(
    db: AsyncSession = Depends(get_db),
    _: Employee = Depends(get_manager_user),
):
    data = await _fetch_matrix_data(db)

    wb = Workbook()
    ws = wb.active
    ws.title = "Матрица компетенций"

    headers = [
        "Сотрудник",
        "Email",
        "Навык",
        "Категория",
        "Самооценка",
        "Оценка руководителя",
        "Цель",
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="2C3E50", end_color="2C3E50", fill_type="solid"
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, r in enumerate(data, 2):
        ws.cell(row=i, column=1, value=r["employee_name"])
        ws.cell(row=i, column=2, value=r["email"])
        ws.cell(row=i, column=3, value=r["skill_name"])
        ws.cell(row=i, column=4, value=r["category"])

        for col, key in [(5, "self_level"), (6, "manager_level")]:
            val = r[key]
            cell = ws.cell(row=i, column=col, value=val or "")
            if val and val in LEVEL_COLORS:
                cell.fill = PatternFill(
                    start_color=LEVEL_COLORS[val],
                    end_color=LEVEL_COLORS[val],
                    fill_type="solid",
                )

        ws.cell(row=i, column=7, value=r["target_level"] or "")

    for col in range(1, 8):
        ws.column_dimensions[chr(64 + col)].width = 22

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=expertise_matrix.xlsx"},
    )


async def _fetch_matrix_data(db: AsyncSession) -> list[dict]:
    result = await db.execute(
        select(Assessment).options(
            selectinload(Assessment.employee), selectinload(Assessment.skill)
        )
    )
    assessments = result.scalars().all()

    rows = []
    for a in assessments:
        rows.append(
            {
                "employee_name": a.employee.full_name,
                "email": a.employee.email,
                "skill_name": a.skill.name,
                "category": a.skill.category,
                "self_level": a.self_level,
                "manager_level": a.manager_level,
                "target_level": a.target_level,
            }
        )
    return sorted(
        rows, key=lambda r: (r["employee_name"], r["category"], r["skill_name"])
    )
